"""Excel I/O: read pending order rows, write outcomes back per line item.

Two rules shape this module.

Non-destructive
    The operators' source file is never written to. ``prepare_working_copy``
    copies it, and every write lands on that copy. A run can be thrown away by
    deleting one file.

Line-item granularity
    The source sheet is order-level - one row per order, with several product
    URLs crammed into a single cell. The spec requires an outcome per SKU, so a
    ``Line Items`` sheet is maintained alongside the original, holding one row
    per SKU with a back-pointer to the order row it came from. The order row
    still gets a rolled-up summary so the sheet stays usable at a glance, but
    the line-item sheet is the record of truth.
"""

from __future__ import annotations

import datetime as dt
import shutil
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    PENDING_STATUS_VALUES,
    LineItem,
    Outcome,
    Platform,
    ReturnStatus,
    TaskStatus,
)

class WorkbookError(Exception):
    """The workbook cannot be interpreted - wrong sheet, or unusable columns."""


ORDERS_SHEET_DEFAULT = "Sheet1"
LINE_ITEMS_SHEET = "Line Items"

#: Without these the sheet cannot be interpreted at all: no way to find the
#: order, know which products it holds, tell whether it is still pending, or
#: know which site to drive.
REQUIRED_COLUMNS = ("Order Id", "Product Link", "Status", "Platform")

#: Normalised header key -> the canonical name the rest of the code uses.
#: Keys are lowercase and stripped of spaces and punctuation, so "Order ID",
#: "order_id" and "OrderId" all land on the same entry.
COLUMN_ALIASES = {
    "orderid": "Order Id",
    "orderno": "Order Id",
    "ordernumber": "Order Id",
    "productlink": "Product Link",
    "productlinks": "Product Link",
    "producturl": "Product Link",
    "producturls": "Product Link",
    "product": "Product Link",
    "status": "Status",
    "taskstatus": "Status",
    "platform": "Platform",
    "site": "Platform",
    "marketplace": "Platform",
    "returnwindow": "Return Window",
    "window": "Return Window",
    "deliverydate": "Delivery date",
    "delivered": "Delivery date",
    "deliveredon": "Delivery date",
    "orderdate": "Order date",
    "ordereddate": "Order date",
    "noofproduct": "No of Product",
    "noofproducts": "No of Product",
    "quantity": "No of Product",
    "qty": "No of Product",
    "amount": "Amount",
    "ordertotal": "Amount",
    "total": "Amount",
    "refundid": "Refund ID",
    "returnid": "Refund ID",
    "returnstatus": "Return Status",
    "refundamount": "Refund Amount",
    "timestamp": "Timestamp",
    "log": "Log",
    "address": "Address",
    "contactnumber": "Contact Number",
    "phone": "Contact Number",
}

#: Columns of the generated per-line-item sheet, in order.
LINE_ITEM_COLUMNS = [
    "Order ID",
    "Platform",
    "Item #",
    "Product / SKU",
    "Product",
    "Product Link",
    "Return Window",
    "Delivery Date",
    "Ordered?",
    "Return ID",
    "Return Status",
    "Detail",
    "Refund Amount",
    "Task Status",
    "Timestamp",
    "Log",
    "Source Row",
]

#: Source-sheet columns the agent writes into on the order row.
ORDER_AGENT_COLUMNS = [
    "Refund ID",
    "Return Status",
    "Refund Amount",
    "Timestamp",
    "Log",
    "Status",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")
_DONE_FILL = PatternFill("solid", fgColor="E2EFDA")
_SKIP_FILL = PatternFill("solid", fgColor="EDEDED")


#: Order-sheet columns written for a run that discovered its own work. They are
#: the input format's own columns, so a discovered result workbook can be read
#: straight back in as the input to a later run.
DISCOVERED_ORDER_COLUMNS = [
    "Order Id",
    "Platform",
    "Product Link",
    "No of Product",
    "Delivery date",
    "Return Window",
    "Status",
    "Refund ID",
    "Return Status",
    "Refund Amount",
    "Timestamp",
    "Log",
]


def prepare_working_copy(source: Path, dest: Path) -> Path:
    """Copy the source workbook to ``dest`` so the original is never modified."""
    source, dest = Path(source), Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    if source.resolve() == dest.resolve():
        raise ValueError("Working copy must differ from the source workbook.")
    shutil.copy2(source, dest)
    return dest


def create_results_workbook(dest: Path, orders_sheet: str = ORDERS_SHEET_DEFAULT) -> Path:
    """Create an empty results workbook for a run that has no input sheet.

    A discovery run has nothing to copy, so the output file is built here
    instead - with the *input* format's own columns, so the file it produces can
    be fed straight back in as the input to a later run.
    """
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = orders_sheet
    for col, name in enumerate(DISCOVERED_ORDER_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    sheet.freeze_panes = "A2"
    book.save(dest)
    return dest


class ReturnsWorkbook:
    """Reads order rows and writes per-line-item outcomes into one workbook."""

    def __init__(self, path: Path, orders_sheet: str = ORDERS_SHEET_DEFAULT):
        self.path = Path(path)
        self.orders_sheet_name = orders_sheet
        # data_only=True resolves any formula cells to their cached values for
        # reading; the write pass reloads without it so formulas survive.
        self._values = openpyxl.load_workbook(self.path, data_only=True)
        if orders_sheet not in self._values.sheetnames:
            # A raw KeyError traceback tells an operator nothing useful; naming
            # the tabs that do exist tells them exactly what to pass to --sheet.
            raise WorkbookError(
                f"This workbook has no sheet named {orders_sheet!r}. "
                f"Sheets present: {', '.join(self._values.sheetnames)}. "
                "Pass the right one with --sheet."
            )
        self._sheet = self._values[orders_sheet]
        self.headers = self._read_headers(self._sheet)

    @staticmethod
    def _normalise(name: object) -> str:
        """Fold a header to a comparison key: lowercase, no spaces or punctuation."""
        return "".join(ch for ch in str(name or "").lower() if ch.isalnum())

    @classmethod
    def _read_headers(cls, sheet) -> dict[str, int]:
        """Map header text to its 1-based column index, tolerating spelling drift.

        Hand-maintained sheets rename columns: ``Order Id`` becomes ``OrderID``,
        ``Product Link`` becomes ``Product URL``. Matching header text exactly
        meant a renamed column silently read as absent, and the agent reported
        nothing to do rather than saying it could not understand the sheet - the
        worst kind of failure, because it looks like success.

        So headers are matched on a normalised key and mapped back to the
        canonical name the rest of the code expects. Unrecognised columns are
        kept under their original name so nothing is lost.
        """
        headers: dict[str, int] = {}
        for idx, cell in enumerate(sheet[1], start=1):
            raw = cell.value
            if raw is None or not str(raw).strip():
                continue
            key = cls._normalise(raw)
            headers[COLUMN_ALIASES.get(key, str(raw).strip())] = idx
        return headers

    @property
    def missing_columns(self) -> list[str]:
        """Required columns this sheet does not provide, under any known spelling."""
        return [c for c in REQUIRED_COLUMNS if c not in self.headers]

    def _cell_value(self, row: int, header: str):
        col = self.headers.get(header)
        return self._sheet.cell(row=row, column=col).value if col else None

    def order_rows(self) -> list[tuple[int, dict]]:
        """Every populated order row as ``(row_number, {header: value})``."""
        rows: list[tuple[int, dict]] = []
        for row in range(2, self._sheet.max_row + 1):
            record = {h: self._cell_value(row, h) for h in self.headers}
            if not str(record.get("Order Id") or "").strip():
                continue
            rows.append((row, record))
        return rows

    def pending_order_rows(self) -> list[tuple[int, dict]]:
        """Order rows whose ``Status`` marks them as not yet processed."""
        pending = []
        for row, record in self.order_rows():
            status = str(record.get("Status") or "").strip().lower()
            if status in PENDING_STATUS_VALUES:
                pending.append((row, record))
        return pending

    # -------------------------------------------------------------- discovery

    def record_discovered_orders(self, items: Iterable[LineItem]) -> int:
        """Give discovered line items a home row, and stamp it onto each one.

        A discovery run has no source sheet, so nothing downstream has a row to
        point at - and the roll-up, the ``Source Row`` back-pointer and resuming
        all key on one. Writing the orders into the sheet first, then stamping
        ``source_row``, means every path after this is the same code the
        spreadsheet run uses. Re-running against the same file updates the
        existing row for an order rather than appending a second one.
        """
        items = list(items)
        if not items:
            return 0

        book = openpyxl.load_workbook(self.path)
        sheet = book[self.orders_sheet_name]

        existing: dict[str, int] = {}
        order_col = self.headers.get("Order Id")
        if order_col:
            for row in range(2, sheet.max_row + 1):
                value = sheet.cell(row=row, column=order_col).value
                if value:
                    existing[str(value).strip()] = row

        groups: dict[str, list[LineItem]] = {}
        for item in items:
            groups.setdefault(item.order_id, []).append(item)

        added = 0
        for order_id, group in groups.items():
            row = existing.get(order_id)
            if row is None:
                row = max(sheet.max_row + 1, 2)
                existing[order_id] = row
                added += 1
            head = group[0]
            payload = {
                "Order Id": order_id,
                "Platform": head.platform.value if head.platform else "Unknown",
                "Product Link": "\n".join(i.product_url for i in group),
                "No of Product": len(group),
                "Delivery date": head.delivery_date.isoformat() if head.delivery_date else "",
                "Return Window": (
                    f"{head.return_window_days} Days" if head.return_window_days else ""
                ),
                "Status": "To Do",
            }
            for header, value in payload.items():
                col = self.headers.get(header)
                if col is None:
                    continue
                cell = sheet.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=(header == "Product Link"))
            for item in group:
                item.source_row = row

        self._autosize(sheet)
        book.save(self.path)
        self._reload()
        return added

    def _reload(self) -> None:
        """Re-read the on-disk state after this object wrote to it."""
        self._values = openpyxl.load_workbook(self.path, data_only=True)
        self._sheet = self._values[self.orders_sheet_name]
        self.headers = self._read_headers(self._sheet)

    # ---------------------------------------------------------------- writing

    def write_results(
        self,
        results: Iterable[tuple[LineItem, Outcome]],
        *,
        dry_run_marker: bool = False,
    ) -> dict:
        """Persist outcomes: per line item, then rolled up to the order row."""
        results = list(results)
        book = openpyxl.load_workbook(self.path)
        orders = book[self.orders_sheet_name]
        line_sheet = self._ensure_line_items_sheet(book)
        index = self._index_line_items(line_sheet)

        written = 0
        for item, outcome in results:
            self._write_line_item(line_sheet, index, item, outcome, dry_run_marker)
            written += 1

        by_row: dict[int, list[tuple[LineItem, Outcome]]] = {}
        for item, outcome in results:
            by_row.setdefault(item.source_row, []).append((item, outcome))

        for source_row, group in by_row.items():
            if source_row < 2:
                # A discovered item that never got recorded into the order sheet
                # - there is no row to roll up onto. The line-item sheet above
                # still holds it, so nothing is lost.
                continue
            self._roll_up_order_row(orders, source_row, group, dry_run_marker)

        self._autosize(line_sheet)
        book.save(self.path)
        return {"line_items_written": written, "order_rows_touched": len(by_row)}

    def _ensure_line_items_sheet(self, book):
        if LINE_ITEMS_SHEET in book.sheetnames:
            sheet = book[LINE_ITEMS_SHEET]
        else:
            sheet = book.create_sheet(LINE_ITEMS_SHEET)
        existing = [c.value for c in sheet[1]] if sheet.max_row >= 1 else []
        if [v for v in existing if v] != LINE_ITEM_COLUMNS:
            for col, name in enumerate(LINE_ITEM_COLUMNS, start=1):
                cell = sheet.cell(row=1, column=col, value=name)
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(vertical="center")
            sheet.freeze_panes = "A2"
        return sheet

    @staticmethod
    def _index_line_items(sheet) -> dict[tuple[str, str], int]:
        """Map ``(order_id, sku)`` to its row so re-runs update in place."""
        index: dict[tuple[str, str], int] = {}
        order_col = LINE_ITEM_COLUMNS.index("Order ID") + 1
        sku_col = LINE_ITEM_COLUMNS.index("Product / SKU") + 1
        for row in range(2, sheet.max_row + 1):
            order_id = sheet.cell(row=row, column=order_col).value
            sku = sheet.cell(row=row, column=sku_col).value
            if order_id:
                index[(str(order_id).strip(), str(sku or "").strip())] = row
        return index

    def _write_line_item(self, sheet, index, item: LineItem, outcome: Outcome, dry: bool):
        key = (item.order_id, item.sku)
        row = index.get(key)
        if row is None:
            row = max([1, *index.values()]) + 1 if index else 2
            row = max(row, sheet.max_row + 1, 2)
            index[key] = row

        window = f"{item.return_window_days} Days" if item.return_window_days else "Unknown"
        delivery = item.delivery_date.isoformat() if item.delivery_date else "Unknown"
        if item.delivery_date_is_approx and item.delivery_date:
            delivery += " (approx)"

        log = outcome.log
        if dry:
            log = f"[DRY RUN - no return submitted] {log}"
        if item.parse_notes:
            log = " | ".join([log, *item.parse_notes]) if log else " | ".join(item.parse_notes)

        values = {
            "Order ID": item.order_id,
            "Platform": item.platform.value if item.platform else "Unknown",
            "Item #": item.item_index + 1,
            "Product / SKU": item.sku,
            "Product": item.title_hint,
            "Product Link": item.product_url,
            "Return Window": window,
            "Delivery Date": delivery,
            "Ordered?": "Yes" if item.ordered else "No (NA)",
            "Return ID": outcome.return_id or "N/A",
            # Spec-conformant vocabulary here; full precision in Detail.
            "Return Status": outcome.spec_status,
            "Detail": outcome.status.value,
            "Refund Amount": outcome.refund_cell,
            "Task Status": outcome.task_status.value,
            "Timestamp": (outcome.timestamp or dt.datetime.now()).isoformat(timespec="seconds"),
            "Log": log,
            "Source Row": item.source_row,
        }

        fill = None
        if outcome.task_status is TaskStatus.NEEDS_REVIEW:
            fill = _REVIEW_FILL
        elif outcome.status is ReturnStatus.NOT_ORDERED:
            fill = _SKIP_FILL
        elif outcome.task_status is TaskStatus.DONE:
            fill = _DONE_FILL

        for col, name in enumerate(LINE_ITEM_COLUMNS, start=1):
            cell = sheet.cell(row=row, column=col, value=values[name])
            cell.alignment = Alignment(vertical="top", wrap_text=(name == "Log"))
            if fill:
                cell.fill = fill

    def _roll_up_order_row(self, sheet, source_row: int, group, dry: bool):
        """Summarise a group of line items onto their originating order row.

        The order row is only marked Done when every one of its line items has
        reached a final state; if any item still needs a human, the whole row is
        flagged for review so nothing is silently dropped.
        """
        actionable = [(i, o) for i, o in group if i.ordered]
        statuses = [o.status for _, o in actionable]
        needs_review = any(o.task_status is TaskStatus.NEEDS_REVIEW for _, o in actionable)
        all_final = all(s.is_final for s in statuses) if statuses else False

        return_ids = [o.return_id for _, o in actionable if o.return_id and o.return_id != "N/A"]
        refunds = [o.refund_amount for _, o in actionable if o.refund_amount is not None]

        if any(s is ReturnStatus.PLANNED for s in statuses):
            # Offline planning only: nothing was attempted, so leave the row
            # Pending rather than claiming a result we never observed.
            row_status = TaskStatus.PENDING.value
        elif needs_review or not all_final:
            row_status = TaskStatus.NEEDS_REVIEW.value
        else:
            row_status = TaskStatus.DONE.value

        # The order row speaks the spec vocabulary too; per-item precision lives
        # on the Line Items sheet.
        counts: dict[str, int] = {}
        for status in statuses:
            label = status.spec_status or status.value
            counts[label] = counts.get(label, 0) + 1
        summary = ", ".join(f"{name} x{n}" for name, n in counts.items()) or "No actionable items"

        skipped = len(group) - len(actionable)
        log_parts = [f"{len(actionable)} line item(s) processed: {summary}."]
        if skipped:
            log_parts.append(f"{skipped} link(s) marked NA in the source cell were not ordered.")
        log_parts.append(f"Per-item detail on the '{LINE_ITEMS_SHEET}' sheet.")
        if dry:
            log_parts.insert(0, "[DRY RUN - no return submitted]")

        payload = {
            "Refund ID": "; ".join(return_ids) if return_ids else "N/A",
            "Return Status": summary,
            "Refund Amount": sum(refunds) if refunds else "N/A",
            "Timestamp": dt.datetime.now().isoformat(timespec="seconds"),
            "Log": " ".join(log_parts),
            "Status": row_status,
        }

        for header, value in payload.items():
            col = self.headers.get(header)
            if col is None:
                continue
            cell = sheet.cell(row=source_row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(header == "Log"))

    @staticmethod
    def _autosize(sheet, max_width: int = 60):
        for col in range(1, sheet.max_column + 1):
            longest = 0
            for row in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row, column=col).value
                if value is not None:
                    longest = max(longest, min(len(str(value)), max_width))
            sheet.column_dimensions[get_column_letter(col)].width = max(10, longest + 2)


def existing_outcomes(path: Path) -> dict[tuple[str, str], str]:
    """Read back already-recorded line-item statuses, for resuming a run."""
    if not Path(path).exists():
        return {}
    book = openpyxl.load_workbook(path, data_only=True)
    if LINE_ITEMS_SHEET not in book.sheetnames:
        return {}
    sheet = book[LINE_ITEMS_SHEET]
    order_col = LINE_ITEM_COLUMNS.index("Order ID") + 1
    sku_col = LINE_ITEM_COLUMNS.index("Product / SKU") + 1
    # Resume decisions need the precise state, not the three-value spec column:
    # "Failed" and "Support Needed" both read as Failed there, but only the first
    # should be re-attempted. Fall back to Return Status for pre-Detail files.
    detail_col = LINE_ITEM_COLUMNS.index("Detail") + 1
    status_col = LINE_ITEM_COLUMNS.index("Return Status") + 1
    out: dict[tuple[str, str], str] = {}
    for row in range(2, sheet.max_row + 1):
        order_id = sheet.cell(row=row, column=order_col).value
        if not order_id:
            continue
        sku = str(sheet.cell(row=row, column=sku_col).value or "").strip()
        status = (
            sheet.cell(row=row, column=detail_col).value
            or sheet.cell(row=row, column=status_col).value
        )
        if status:
            out[(str(order_id).strip(), sku)] = str(status).strip()
    return out
