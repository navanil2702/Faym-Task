"""Write-back tests: line-item granularity and partial-success roll-up."""

from __future__ import annotations

import datetime as dt
import shutil
from pathlib import Path

import openpyxl
import pytest

from faym_returns.models import LineItem, Outcome, Platform, ReturnStatus, TaskStatus
from faym_returns.workbook import (
    LINE_ITEM_COLUMNS,
    LINE_ITEMS_SHEET,
    ReturnsWorkbook,
    prepare_working_copy,
)

SOURCE = Path.home() / "Downloads" / "Faym Status Test Orders.xlsx"

pytestmark = pytest.mark.skipif(
    not SOURCE.exists(), reason="test dataset not present in ~/Downloads"
)


@pytest.fixture()
def book(tmp_path: Path) -> ReturnsWorkbook:
    working = tmp_path / "results.xlsx"
    prepare_working_copy(SOURCE, working)
    return ReturnsWorkbook(working)


def _cell(path: Path, column: str, row: int = 2):
    """Read a Line Items cell by column name, not by a brittle index."""
    sheet = openpyxl.load_workbook(path)[LINE_ITEMS_SHEET]
    return sheet.cell(row=row, column=LINE_ITEM_COLUMNS.index(column) + 1).value


def _item(sku: str, row: int = 6, index: int = 0, ordered: bool = True) -> LineItem:
    return LineItem(
        source_row=row,
        item_index=index,
        order_id="OD337974610559997100",
        platform=Platform.FLIPKART,
        sku=sku,
        product_url=f"https://www.flipkart.com/x/p/itm1?pid={sku}",
        title_hint="test product",
        delivery_date=dt.date(2026, 7, 6),
        return_window_days=10,
        ordered=ordered,
    )


def test_source_workbook_is_never_modified(tmp_path: Path):
    before = SOURCE.read_bytes()
    working = tmp_path / "copy.xlsx"
    prepare_working_copy(SOURCE, working)
    book = ReturnsWorkbook(working)
    book.write_results([(_item("AAA"), Outcome(ReturnStatus.PLACED, "RET1").stamp())])
    assert SOURCE.read_bytes() == before


def test_working_copy_must_differ_from_source():
    with pytest.raises(ValueError):
        prepare_working_copy(SOURCE, SOURCE)


def test_reads_pending_order_rows():
    working = SOURCE  # read-only use
    book = ReturnsWorkbook(working)
    rows = book.pending_order_rows()
    assert len(rows) == 7, "the dataset has 7 populated order rows, all Pending"
    assert all(str(r["Status"]).strip().lower() == "pending" for _, r in rows)


def test_write_results_creates_one_row_per_line_item(book: ReturnsWorkbook):
    results = [
        (_item("AAA", index=0), Outcome(ReturnStatus.PLACED, "RET-A", 284.0).stamp()),
        (_item("BBB", index=1), Outcome(ReturnStatus.OUT_OF_WINDOW).stamp()),
        (_item("CCC", index=2), Outcome(ReturnStatus.PLACED, "RET-C", 350.0).stamp()),
    ]
    stats = book.write_results(results)
    assert stats["line_items_written"] == 3

    sheet = openpyxl.load_workbook(book.path)[LINE_ITEMS_SHEET]
    skus = [sheet.cell(row=r, column=4).value for r in range(2, sheet.max_row + 1)]
    assert skus == ["AAA", "BBB", "CCC"]


def test_rerun_updates_in_place_rather_than_appending(book: ReturnsWorkbook):
    book.write_results([(_item("AAA"), Outcome(ReturnStatus.FAILED).stamp())])
    book.write_results([(_item("AAA"), Outcome(ReturnStatus.PLACED, "RET-A").stamp())])

    sheet = openpyxl.load_workbook(book.path)[LINE_ITEMS_SHEET]
    assert sheet.max_row == 2, "the same SKU must not produce a second row"
    assert sheet.cell(row=2, column=11).value == ReturnStatus.PLACED.value


def test_partial_success_flags_the_order_for_review(book: ReturnsWorkbook):
    """Two items succeed, one needs a human -> the order row is flagged, and
    the successful returns are still recorded."""
    results = [
        (_item("AAA", index=0), Outcome(ReturnStatus.PLACED, "RET-A", 284.0).stamp()),
        (_item("BBB", index=1), Outcome(ReturnStatus.PLACED, "RET-B", 350.0).stamp()),
        (_item("CCC", index=2), Outcome(ReturnStatus.SUPPORT_NEEDED).stamp()),
    ]
    book.write_results(results)

    orders = openpyxl.load_workbook(book.path)["Sheet1"]
    headers = {c.value: c.column for c in orders[1] if c.value}
    status = orders.cell(row=6, column=headers["Status"]).value
    refund_ids = orders.cell(row=6, column=headers["Refund ID"]).value
    refund_total = orders.cell(row=6, column=headers["Refund Amount"]).value

    assert status == TaskStatus.NEEDS_REVIEW.value
    assert "RET-A" in refund_ids and "RET-B" in refund_ids
    assert refund_total == 634.0


def test_order_marked_done_only_when_all_items_final(book: ReturnsWorkbook):
    results = [
        (_item("AAA", index=0), Outcome(ReturnStatus.PLACED, "RET-A", 100.0).stamp()),
        (_item("BBB", index=1), Outcome(ReturnStatus.OUT_OF_WINDOW).stamp()),
    ]
    book.write_results(results)
    orders = openpyxl.load_workbook(book.path)["Sheet1"]
    headers = {c.value: c.column for c in orders[1] if c.value}
    assert orders.cell(row=6, column=headers["Status"]).value == TaskStatus.DONE.value


def test_failed_item_keeps_order_out_of_done(book: ReturnsWorkbook):
    results = [
        (_item("AAA", index=0), Outcome(ReturnStatus.PLACED, "RET-A").stamp()),
        (_item("BBB", index=1), Outcome(ReturnStatus.FAILED).stamp()),
    ]
    book.write_results(results)
    orders = openpyxl.load_workbook(book.path)["Sheet1"]
    headers = {c.value: c.column for c in orders[1] if c.value}
    assert orders.cell(row=6, column=headers["Status"]).value == TaskStatus.NEEDS_REVIEW.value


def test_na_items_are_excluded_from_the_roll_up(book: ReturnsWorkbook):
    """A never-ordered link must not drag the order row into review."""
    results = [
        (_item("AAA", index=0), Outcome(ReturnStatus.PLACED, "RET-A", 100.0).stamp()),
        (
            _item("BBB", index=1, ordered=False),
            Outcome(ReturnStatus.NOT_ORDERED).stamp(),
        ),
    ]
    book.write_results(results)
    orders = openpyxl.load_workbook(book.path)["Sheet1"]
    headers = {c.value: c.column for c in orders[1] if c.value}
    assert orders.cell(row=6, column=headers["Status"]).value == TaskStatus.DONE.value
    log = orders.cell(row=6, column=headers["Log"]).value
    assert "marked NA" in log


def test_dry_run_is_marked_in_the_log(book: ReturnsWorkbook):
    book.write_results(
        [(_item("AAA"), Outcome(ReturnStatus.PLACED, "RET-A").stamp())],
        dry_run_marker=True,
    )
    assert "DRY RUN" in _cell(book.path, "Log")


def test_refund_amount_is_na_when_platform_showed_none(book: ReturnsWorkbook):
    book.write_results([(_item("AAA"), Outcome(ReturnStatus.OUT_OF_WINDOW).stamp())])
    assert _cell(book.path, "Refund Amount") == "N/A"


# ------------------------------------------------- spec-conformant status column


@pytest.mark.parametrize(
    "status,expected",
    [
        (ReturnStatus.PLACED, "Placed"),
        (ReturnStatus.ALREADY_REFUNDED, "Placed"),
        (ReturnStatus.OUT_OF_WINDOW, "Out of window"),
        (ReturnStatus.NOT_DELIVERED, "Failed"),
        (ReturnStatus.SUPPORT_NEEDED, "Failed"),
        (ReturnStatus.ITEM_NOT_FOUND, "Failed"),
        (ReturnStatus.NOT_ORDERED, "Failed"),
        (ReturnStatus.FAILED, "Failed"),
        # Nothing attempted: asserting a failure would be untrue.
        (ReturnStatus.PLANNED, None),
    ],
)
def test_return_status_column_uses_only_the_three_spec_values(
    book: ReturnsWorkbook, status: ReturnStatus, expected: str
):
    book.write_results([(_item("AAA"), Outcome(status).stamp())])
    assert _cell(book.path, "Return Status") == expected


def test_detail_column_preserves_the_precise_state(book: ReturnsWorkbook):
    """Collapsing to three values must not lose what actually happened."""
    book.write_results(
        [(_item("AAA"), Outcome(ReturnStatus.ALREADY_REFUNDED, "CR123").stamp())]
    )
    assert _cell(book.path, "Return Status") == "Placed"
    assert _cell(book.path, "Detail") == "Already Cancelled & Refunded"


def test_resume_reads_the_detail_column_not_the_spec_column(book: ReturnsWorkbook):
    """Support Needed and Failed both read as "Failed" in the spec column, but
    only Failed should be re-attempted - so resume must use Detail."""
    from faym_returns.workbook import existing_outcomes

    book.write_results([(_item("AAA"), Outcome(ReturnStatus.SUPPORT_NEEDED).stamp())])
    recorded = existing_outcomes(book.path)
    assert recorded[("OD337974610559997100", "AAA")] == "Support Needed"


# --------------------------------------------- tolerating other people's sheets


def _sheet(tmp_path: Path, headers: list, rows: list, title: str = "Sheet1") -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    for r in rows:
        ws.append(r)
    p = tmp_path / "other.xlsx"
    wb.save(p)
    return p


BASE = ["Order Id", "Product Link", "Status", "Platform", "Return Window", "Delivery date"]
ROW = ["OD1", "https://www.flipkart.com/x/p/itm1?pid=ABC123", "Pending", "Flipkart",
       "10 Days", dt.datetime(2026, 7, 20)]


def test_header_matching_ignores_case_spacing_and_punctuation():
    """Hand-maintained sheets drift: OrderID, order_id, Order ID all mean one thing."""
    from faym_returns.workbook import ReturnsWorkbook as W
    for variant in ["Order Id", "OrderID", "order_id", "ORDER ID", "Order  Id"]:
        assert W._normalise(variant) == "orderid"


def test_renamed_columns_are_still_found(tmp_path: Path):
    """A sheet using OrderID and Product URL must still be readable."""
    p = _sheet(tmp_path, ["OrderID", "Product URL", "Status", "Site",
                          "Window", "Delivered On"], [ROW])
    book = ReturnsWorkbook(p)
    assert book.missing_columns == []
    assert len(book.pending_order_rows()) == 1


def test_unusable_columns_are_reported_not_silently_empty(tmp_path: Path):
    """The dangerous failure: finding nothing looks identical to having nothing."""
    p = _sheet(tmp_path, ["Customer", "Notes", "Total"], [["x", "y", 1]])
    book = ReturnsWorkbook(p)
    assert set(book.missing_columns) == {"Order Id", "Product Link", "Status", "Platform"}


def test_a_missing_sheet_names_the_ones_that_exist(tmp_path: Path):
    from faym_returns.workbook import WorkbookError

    p = _sheet(tmp_path, BASE, [ROW], title="Orders")
    with pytest.raises(WorkbookError, match="Sheets present: Orders"):
        ReturnsWorkbook(p)
    assert ReturnsWorkbook(p, "Orders").missing_columns == []
