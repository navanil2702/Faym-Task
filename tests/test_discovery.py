"""Tests for the input path that has no spreadsheet.

A live run builds its work list by signing in and walking My Orders. Nothing
here drives a browser: the page is stubbed, so what is under test is the part
that turns order cards into line items, and the part that gives those line items
somewhere to be written.
"""

from __future__ import annotations

import datetime as dt

import openpyxl
import pytest

from faym_returns import cli, eligibility
from faym_returns.models import LineItem, Outcome, Platform, ReturnStatus
from faym_returns.orchestrator import Orchestrator, RunOptions, RunReport
from faym_returns.platforms.amazon import AmazonAdapter
from faym_returns.platforms.flipkart import FlipkartAdapter
from faym_returns.workbook import (
    LINE_ITEMS_SHEET,
    ReturnsWorkbook,
    create_results_workbook,
)

TODAY = dt.date(2026, 7, 20)


class FakeAnchors:
    def __init__(self, hrefs):
        self._hrefs = hrefs

    def evaluate_all(self, _script):
        return list(self._hrefs)


class FakeCard:
    """The slice of a Playwright Locator that card parsing actually touches."""

    def __init__(self, text: str, hrefs=()):
        self._text = text
        self._hrefs = list(hrefs)

    def inner_text(self, timeout=None):
        return self._text

    def locator(self, selector):
        assert selector == "a"
        return FakeAnchors(self._hrefs)


@pytest.fixture
def flipkart():
    # The card helpers never touch the session; only the selector file matters.
    return FlipkartAdapter(session=None)


@pytest.fixture
def amazon():
    return AmazonAdapter(session=None)


def _card_items(adapter, card, *, within_days=30):
    return adapter._line_items_from_card(card, today=TODAY, within_days=within_days)


# ------------------------------------------------------- reading an order card


def test_a_delivered_card_becomes_one_line_item_per_product(flipkart):
    card = FakeCard(
        "Order ID - OD337915012166989100\nDelivered on 18 Jul 2026\nReturn",
        [
            "https://www.flipkart.com/some-shirt/p/itmabc123?pid=TSHG9FQZSSAUGKUP",
            "https://www.flipkart.com/some-jeans/p/itmdef456?pid=JEAH87B2GRCCS3DZ",
            "https://www.flipkart.com/account/orders",  # not a product link
        ],
    )
    items = _card_items(flipkart, card)

    assert [i.sku for i in items] == ["TSHG9FQZSSAUGKUP", "JEAH87B2GRCCS3DZ"]
    assert all(i.order_id == "OD337915012166989100" for i in items)
    assert all(i.platform is Platform.FLIPKART for i in items)
    assert [i.item_index for i in items] == [0, 1]
    assert items[0].delivery_date == dt.date(2026, 7, 18)


def test_the_same_product_linked_twice_is_one_line_item(flipkart):
    """Order cards link the same product from its image and its title."""
    url = "https://www.flipkart.com/a-shirt/p/itmabc?pid=TSHG9FQZSSAUGKUP"
    card = FakeCard("Order ID - OD1234567890123456\nDelivered\nReturn", [url, url])
    assert len(_card_items(flipkart, card)) == 1


def test_a_discovered_item_carries_no_source_row_yet(flipkart):
    card = FakeCard(
        "OD1234567890123456 Delivered Return",
        ["https://www.flipkart.com/x/p/itmabc?pid=AAA111"],
    )
    item = _card_items(flipkart, card)[0]
    assert item.source_row == 0
    assert item.ordered is True
    assert any("Discovered" in note for note in item.parse_notes)


@pytest.mark.parametrize(
    "copy",
    [
        "OD1234567890123456 Refund completed",
        "OD1234567890123456 Cancelled",
        "OD1234567890123456 Delivered Return request placed",
        "OD1234567890123456 Delivered Return window closed",
    ],
)
def test_already_settled_cards_are_left_alone(flipkart, copy):
    """No point opening an order the list already reports as done with."""
    card = FakeCard(copy, ["https://www.flipkart.com/x/p/itmabc?pid=AAA111"])
    assert not _card_items(flipkart, card)


def test_an_undelivered_card_is_not_picked_up(flipkart):
    card = FakeCard(
        "OD1234567890123456 Out for delivery",
        ["https://www.flipkart.com/x/p/itmabc?pid=AAA111"],
    )
    assert not _card_items(flipkart, card)


def test_a_card_without_an_order_id_is_skipped(flipkart):
    card = FakeCard("Delivered Return", ["https://www.flipkart.com/x/p/itmabc?pid=AAA111"])
    assert _card_items(flipkart, card) is None


def test_orders_older_than_the_lookback_are_skipped(flipkart):
    card = FakeCard(
        "OD1234567890123456 Delivered on 1 Jun 2026 Return",
        ["https://www.flipkart.com/x/p/itmabc?pid=AAA111"],
    )
    assert not _card_items(flipkart, card, within_days=30)
    assert _card_items(flipkart, card, within_days=90)


def test_a_stated_return_window_feeds_the_eligibility_check(flipkart):
    """The spec's `Return window` field, sourced from the page instead of a cell."""
    card = FakeCard(
        "OD1234567890123456 Delivered on 18 Jul 2026\n10 days return policy",
        ["https://www.flipkart.com/x/p/itmabc?pid=AAA111"],
    )
    item = _card_items(flipkart, card)[0]

    assert item.return_window_days == 10
    # Delivered 18 Jul + 10 days, checked on 1 Sep: clearly expired, and now the
    # local pre-filter can say so instead of deferring.
    verdict = eligibility.check(item, today=dt.date(2026, 9, 1))
    assert verdict.eligible is False


def test_an_unstated_return_window_says_so_rather_than_going_blank(flipkart):
    card = FakeCard(
        "OD1234567890123456 Delivered on 18 Jul 2026 Return",
        ["https://www.flipkart.com/x/p/itmabc?pid=AAA111"],
    )
    item = _card_items(flipkart, card)[0]

    assert item.return_window_days is None
    assert any("not stated" in note for note in item.parse_notes)
    # Unknown is not expired: the platform is left to make the call.
    assert eligibility.check(item, today=dt.date(2027, 1, 1)).eligible is True


@pytest.mark.parametrize(
    "copy, expected",
    [
        ("7 days return", 7),
        ("10 Days Easy Return", 10),
        ("Return window: 14 days", 14),
        ("Returnable within 30 days", 30),
        ("No return available", None),
    ],
)
def test_return_window_copy_shapes(flipkart, copy, expected):
    assert flipkart._return_window_from(copy) == expected


def test_a_card_with_no_readable_date_is_kept(flipkart):
    """Unknown is not the same as expired - the platform decides."""
    card = FakeCard(
        "OD1234567890123456 Delivered Return",
        ["https://www.flipkart.com/x/p/itmabc?pid=AAA111"],
    )
    items = _card_items(flipkart, card)
    assert len(items) == 1
    assert items[0].delivery_date is None


def test_amazon_cards_read_their_own_id_and_date_shapes(amazon):
    card = FakeCard(
        "Order # 407-1234567-1234567\nDelivered Jul 18, 2026\nReturn or replace items",
        ["https://www.amazon.in/Some-Product-Name/dp/B0ABCDEFGH/ref=x"],
    )
    items = _card_items(amazon, card)
    assert len(items) == 1
    assert items[0].order_id == "407-1234567-1234567"
    assert items[0].sku == "B0ABCDEFGH"
    assert items[0].delivery_date == dt.date(2026, 7, 18)
    assert items[0].platform is Platform.AMAZON


# ---------------------------------------------------- a workbook with no input


def _discovered(order_id="OD1234567890123456", sku="AAA111", index=0):
    return LineItem(
        source_row=0,
        item_index=index,
        order_id=order_id,
        platform=Platform.FLIPKART,
        sku=sku,
        product_url=f"https://www.flipkart.com/x/p/itmabc?pid={sku}",
        title_hint="a shirt",
        delivery_date=dt.date(2026, 7, 18),
    )


def test_a_results_workbook_is_created_from_nothing(tmp_path):
    path = create_results_workbook(tmp_path / "returns.xlsx")
    book = ReturnsWorkbook(path)
    assert book.missing_columns == []
    assert book.order_rows() == []


def test_discovered_orders_get_a_row_and_a_back_pointer(tmp_path):
    path = create_results_workbook(tmp_path / "returns.xlsx")
    book = ReturnsWorkbook(path)
    items = [_discovered(sku="AAA111", index=0), _discovered(sku="BBB222", index=1)]

    added = book.record_discovered_orders(items)

    assert added == 1  # two line items, one order
    assert {i.source_row for i in items} == {2}
    rows = book.order_rows()
    assert len(rows) == 1
    record = rows[0][1]
    assert record["Order Id"] == "OD1234567890123456"
    assert record["No of Product"] == 2
    assert record["Platform"] == "Flipkart"


def test_rediscovering_an_order_updates_its_row(tmp_path):
    """A second run must not append the same order again."""
    path = create_results_workbook(tmp_path / "returns.xlsx")
    book = ReturnsWorkbook(path)
    book.record_discovered_orders([_discovered()])
    again = [_discovered()]

    assert book.record_discovered_orders(again) == 0
    assert again[0].source_row == 2
    assert len(book.order_rows()) == 1


def test_two_orders_land_on_separate_rows(tmp_path):
    path = create_results_workbook(tmp_path / "returns.xlsx")
    book = ReturnsWorkbook(path)
    items = [_discovered(order_id="OD111", sku="A"), _discovered(order_id="OD222", sku="B")]

    assert book.record_discovered_orders(items) == 2
    assert items[0].source_row != items[1].source_row


def test_outcomes_write_back_exactly_as_they_would_from_a_sheet(tmp_path):
    """Past record_discovered_orders, a discovered item is an ordinary item."""
    path = create_results_workbook(tmp_path / "returns.xlsx")
    book = ReturnsWorkbook(path)
    items = [_discovered()]
    book.record_discovered_orders(items)

    outcome = Outcome(
        status=ReturnStatus.PLACED, return_id="RET12345678", refund_amount=953.0
    ).stamp()
    stats = book.write_results([(items[0], outcome)])

    assert stats == {"line_items_written": 1, "order_rows_touched": 1}
    saved = openpyxl.load_workbook(path)
    assert LINE_ITEMS_SHEET in saved.sheetnames
    line = saved[LINE_ITEMS_SHEET]
    header = [c.value for c in line[1]]
    row = [c.value for c in line[2]]
    assert row[header.index("Return ID")] == "RET12345678"
    assert row[header.index("Return Status")] == "Placed"
    assert row[header.index("Source Row")] == 2


# --------------------------------------------------------------- orchestration


def _orchestrator(path, **overrides):
    options = dict(
        discover=True,
        dry_run=True,
        discover_platforms=(Platform.FLIPKART,),
        today=TODAY,
    )
    options.update(overrides)
    return Orchestrator(ReturnsWorkbook(path), None, RunOptions(**options))


def test_discovery_mode_plans_nothing_from_the_workbook(tmp_path):
    """The output file is an output file: nothing is ever read back as work."""
    path = create_results_workbook(tmp_path / "returns.xlsx")
    book = ReturnsWorkbook(path)
    book.record_discovered_orders([_discovered()])

    to_attempt, decided = _orchestrator(path).plan()

    assert to_attempt == []
    assert decided == []


def test_triage_defers_to_the_platform_when_a_card_gave_no_window(tmp_path):
    path = create_results_workbook(tmp_path / "returns.xlsx")
    item = _discovered()
    item.delivery_date = None

    to_attempt, decided = _orchestrator(path).triage([item])

    assert to_attempt == [item]
    assert decided == []


def test_triage_still_rejects_a_clearly_expired_discovered_item(tmp_path):
    path = create_results_workbook(tmp_path / "returns.xlsx")
    item = _discovered()
    item.delivery_date = dt.date(2026, 5, 1)
    item.return_window_days = 7

    to_attempt, decided = _orchestrator(path).triage([item])

    assert to_attempt == []
    assert decided[0][1].status is ReturnStatus.OUT_OF_WINDOW


def test_the_limit_is_a_total_not_a_per_platform_allowance(tmp_path):
    path = create_results_workbook(tmp_path / "returns.xlsx")
    items = [_discovered(order_id=f"OD{n}", sku=f"SKU{n}") for n in range(5)]

    to_attempt, _ = _orchestrator(path, limit=2).triage(items, limit=2)

    assert len(to_attempt) == 2


class FakeAdapter:
    """An adapter that has already 'walked' the orders page."""

    platform = Platform.FLIPKART

    def __init__(self, items):
        self.items = items
        self.called_with = None

    def discover_returnable(self, **kwargs):
        self.called_with = kwargs
        return list(self.items)


def test_discovered_work_is_recorded_before_any_return_is_attempted(tmp_path):
    """Every discovered item has a row before the agent touches a return flow.

    That ordering is what lets the write-back, the roll-up and resuming work
    identically for both input paths.
    """
    path = create_results_workbook(tmp_path / "returns.xlsx")
    orchestrator = _orchestrator(path)
    report = RunReport()
    expired = _discovered(order_id="OD999", sku="OLD1")
    expired.delivery_date = dt.date(2026, 5, 1)
    expired.return_window_days = 7
    adapter = FakeAdapter([_discovered(), expired])

    orders = orchestrator._discover_on(adapter, report)

    assert list(orders) == ["OD1234567890123456"]
    # Both the attemptable item and the expired one hold a real row.
    assert len(report.planned) == 2
    assert all(item.source_row >= 2 for item in report.planned)
    # The expired one is already settled and never reaches a return flow.
    assert [o.status for _, o in report.results] == [ReturnStatus.OUT_OF_WINDOW]


def test_discovery_passes_its_bounds_to_the_adapter(tmp_path):
    path = create_results_workbook(tmp_path / "returns.xlsx")
    orchestrator = _orchestrator(path, discover_within_days=7, discover_max_orders=3)
    adapter = FakeAdapter([_discovered()])

    orchestrator._discover_on(adapter, RunReport())

    assert adapter.called_with == {
        "max_orders": 3,
        "within_days": 7,
        "today": TODAY,
    }


def test_a_spent_limit_stops_discovery_before_it_walks_a_second_site(tmp_path):
    path = create_results_workbook(tmp_path / "returns.xlsx")
    orchestrator = _orchestrator(path, limit=1)
    report = RunReport(planned=[_discovered()])
    adapter = FakeAdapter([_discovered(order_id="OD222", sku="B")])

    assert orchestrator._discover_on(adapter, report) == {}
    assert adapter.called_with is None  # the page was never walked


# -------------------------------------------------------------- write-back timing


class ExplodingSession:
    """A session that dies the moment the run tries to use it."""

    def __init__(self, *args, **kwargs):
        pass

    def __enter__(self):
        raise KeyboardInterrupt("second Ctrl-C")

    def __exit__(self, *exc):
        return False


def test_a_hard_interrupt_still_writes_what_was_gathered(tmp_path, monkeypatch):
    """A filed return that was never recorded is the expensive failure.

    KeyboardInterrupt is not an Exception, so it escapes the handlers around the
    run loop. The write-back has to happen on the way out regardless.
    """
    import faym_returns.orchestrator as orch_module

    path = create_results_workbook(tmp_path / "returns.xlsx")
    done, pending = _discovered(sku="DONE1"), _discovered(sku="TODO1", index=1)
    book = ReturnsWorkbook(path)
    book.record_discovered_orders([done, pending])

    orchestrator = Orchestrator(book, None, RunOptions(discover=False, today=TODAY))
    outcome = Outcome(status=ReturnStatus.PLACED, return_id="RET99").stamp()
    # One item already has a result; the next one kills the session.
    monkeypatch.setattr(orchestrator, "plan", lambda: ([pending], [(done, outcome)]))
    monkeypatch.setattr(orch_module, "Session", ExplodingSession)

    with pytest.raises(KeyboardInterrupt):
        orchestrator.run()

    saved = openpyxl.load_workbook(path)[LINE_ITEMS_SHEET]
    header = [c.value for c in saved[1]]
    assert saved.max_row == 2, "the outcome was lost on the way out"
    assert saved[2][header.index("Return ID")].value == "RET99"


# ------------------------------------------------------------------------- cli


def test_a_discovery_run_creates_its_own_results_file(tmp_path, monkeypatch):
    """No input file exists, so the CLI has to build the output from scratch."""
    monkeypatch.chdir(tmp_path)

    captured = {}

    class StubOrchestrator:
        def __init__(self, book, *args, **kwargs):
            captured["path"] = book.path

        def run(self):
            return RunReport()

    monkeypatch.setattr(cli, "Orchestrator", StubOrchestrator)
    assert cli.main(["--discover", "--platform", "Flipkart", "--allow-quiet-hours"]) == 0

    created = list(tmp_path.glob("returns-*.xlsx"))
    assert len(created) == 1
    assert captured["path"] == created[0]
    assert ReturnsWorkbook(created[0]).missing_columns == []
